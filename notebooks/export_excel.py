"""
Builds the final Excel workbook summarizing the analysis, with sheets for
Raw Data, Downtime Summary, and Utilization Summary.

Run after the three notebooks (they populate outputs/exports/*.csv):
    python notebooks/export_excel.py
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
EXPORTS = ROOT / "outputs" / "exports"
OUT_PATH = ROOT / "outputs" / "exports" / "healthcare_imaging_summary.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def autosize_and_style(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def main():
    merged = pd.read_csv(EXPORTS / "merged_equipment_metadata.csv")
    downtime_summary = pd.read_csv(EXPORTS / "downtime_summary.csv")
    utilization_summary = pd.read_csv(EXPORTS / "utilization_summary.csv")
    top_causes = pd.read_csv(EXPORTS / "top_downtime_causes.csv")

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        # Raw Data sheet: sample of the merged, cleaned dataset (full history
        # is thousands of rows -- keep the workbook a reasonable size while
        # still giving a representative look at the row-level data).
        merged.sample(n=min(500, len(merged)), random_state=42).sort_values(
            ["machine_id", "date"]
        ).to_excel(writer, sheet_name="Raw Data", index=False)

        downtime_summary.to_excel(writer, sheet_name="Downtime Summary", index=False)
        utilization_summary.to_excel(writer, sheet_name="Utilization Summary", index=False)
        top_causes.to_excel(writer, sheet_name="Top Downtime Causes", index=False)

    from openpyxl import load_workbook

    wb = load_workbook(OUT_PATH)
    for sheet_name in wb.sheetnames:
        autosize_and_style(wb[sheet_name])
    wb.save(OUT_PATH)

    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
